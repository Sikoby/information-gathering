"""Flushed run artifacts (out/<run_id>/) for finished meetings.

The agent writes these files at shutdown (src/persistence.py); the console
container sees them through a read-only ./out bind mount. The JSON is re-read
directly here — importing src.persistence would drag agent code (LiveKit)
into this image. Run ids come from the trusted Redis record, never from the
URL, so no path-traversal guard is needed.
"""

from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# The extractor's catch-all question (findings it couldn't place). Hide it
# from the export when nothing ever landed there.
_CATCH_ALL_QUESTION_ID = "other/q"


def _out_root() -> Path:
    return Path(os.environ.get("CONSOLE_OUT_ROOT", "out"))


def _run_dir(run_id: str) -> Path:
    return _out_root() / run_id


def load_sections(run_id: str) -> list[dict] | None:
    """The flushed Section tree, or None when the agent never flushed."""
    path = _run_dir(run_id) / "tree.json"
    try:
        return json.loads(path.read_text())
    except FileNotFoundError:
        return None
    except (json.JSONDecodeError, OSError) as e:
        logger.warning("unreadable tree.json for run_id={}: {}", run_id, e)
        return None


def load_transcript(run_id: str) -> list[dict] | None:
    """transcript.jsonl as a list of {ts, role, text} lines.

    The file is appended live, so a SIGKILL can truncate the last line —
    malformed lines are skipped rather than failing the whole read.
    """
    path = _run_dir(run_id) / "transcript.jsonl"
    try:
        raw = path.read_text()
    except FileNotFoundError:
        return None
    except OSError as e:
        logger.warning("unreadable transcript for run_id={}: {}", run_id, e)
        return None
    lines: list[dict] = []
    for line in raw.splitlines():
        if not line.strip():
            continue
        try:
            lines.append(json.loads(line))
        except json.JSONDecodeError:
            logger.warning("skipping malformed transcript line for run_id={}", run_id)
    return lines


def load_results(run_id: str) -> dict:
    """The detail-page payload: null fields mean "never flushed"."""
    return {
        "sections": load_sections(run_id),
        "transcript": load_transcript(run_id),
    }


def _fmt_ts(ts: str | None) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromisoformat(ts).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return ts


def build_answers_xlsx(sections: list[dict]) -> bytes:
    """The question→answer sheet. Topic columns mirror the tree depth
    (Section, Subsection, …), then Question / Answer / Detail / Recorded at.
    One row per answer; an unanswered question keeps one row with empty
    answer cells so gaps stay visible.
    """
    children: dict[str | None, list[dict]] = defaultdict(list)
    for s in sections:
        children[s.get("parent_id")].append(s)

    # DFS from the meeting root, collecting each question with the headers
    # of its topic ancestors (root excluded). File order is preserved.
    questions: list[tuple[list[str], dict, list[dict]]] = []

    def walk(section_id: str, path: list[str]) -> None:
        for child in children.get(section_id, []):
            if child["kind"] == "question":
                answers = sorted(
                    (a for a in children.get(child["id"], []) if a["kind"] == "answer"),
                    key=lambda a: a.get("ts") or "",
                )
                if not answers and child["id"] == _CATCH_ALL_QUESTION_ID:
                    continue
                questions.append((path, child, answers))
            elif child["kind"] == "topic":
                walk(child["id"], path + [child["header"]])

    roots = [s for s in sections if s["kind"] == "meeting"] or [
        s for s in sections if s.get("parent_id") is None
    ]
    for root in roots:
        walk(root["id"], [])

    depth = max((len(path) for path, _, _ in questions), default=1)
    topic_headers = ["Section", "Subsection"][:depth] + [
        f"Level {i}" for i in range(3, depth + 1)
    ]
    headers = topic_headers + ["Question", "Answer", "Detail", "Recorded at"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for path, question, answers in questions:
        topic_cells = path + [""] * (depth - len(path))
        if answers:
            for a in answers:
                ws.append(
                    topic_cells
                    + [
                        question["header"],
                        a.get("header") or "",
                        a.get("body") or "",
                        _fmt_ts(a.get("ts")),
                    ]
                )
        else:
            ws.append(topic_cells + [question["header"], "", "", ""])

    widths = [28] * depth + [48, 32, 60, 17]
    wrapped = set(range(depth + 1, depth + 4))  # Question, Answer, Detail
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in wrapped:
                cell.alignment = wrap

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
